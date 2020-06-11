from openpyxl import *
class Ogrenci:
    def __init__(self):
        self.id=0
        self.numara=''
        self.ad_soyad=''
        self.sinif=''
        self.sinavlar=[]

class VeriAl:
    veriler = load_workbook("sinavlar.xlsx")
    sheet=veriler.active
    ogrenci=[]
    n=0
    for x,row in enumerate(sheet.rows):
       if x>3:
            if type(sheet.cell(x+1,1).value) is int and type(sheet.cell(x+1,2).value) is int:
                ogr=Ogrenci()
                ogr.numara=sheet.cell(x+1,2).value
                ogr.ad_soyad=sheet.cell(x+1,3).value
                ogr.sinif=sheet.cell(x+1,9).value
                ogr.sinavlar.append(sheet.cell(x+1,10).value)
                ogr.id=n
                ogrenci.append(ogr)
                n=n+1
                #ders=sheet.cell(x+1,10).value
            elif type(sheet.cell(x+1,9).value) is int:

                ogrenci[n-1].sinavlar.append(sheet.cell(x+1,10).value)
                
  
    for i in ogrenci:
        print(i.ad_soyad)  
        for j in i.sinavlar:
            print(j)          
    # for i,row in enumerate(sheet.iter_cols(min_row=5, min_col=10, max_row=90, max_col=10)):
    #     #for cell in row:
    #      print(row[i].value,i)
    # for row in sheet.rows:
    #     for cell in row:
    #       print(cell.value)
    veriler.close